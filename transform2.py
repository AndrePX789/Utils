from openpyxl import load_workbook
import os
import datetime


try:
    os.remove('output.txt')
except OSError:
    print('New file')

file = open('output.txt', 'w')

temp_table_name = 'solicitantes'
insert_command = f'INSERT INTO {temp_table_name} (nome, endereco, cidade, estado, cep, data_nascimento, data_cadastro) VALUES\n'

workbook = load_workbook('..\\Solicitantes_SeparadosPorInicial/Letra A.xlsx', read_only=True)
sheets_names = workbook.sheetnames

# expecting only 'oferta recomendadas'

sheet = workbook[sheets_names[0]]

file.write('SET @SolicitanteId = 0;\n\n')

for row in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
    nome = str(row[1]).title().strip().replace('\'', "")

    if row[2] is not None:
        endereco = row[2].title().strip().replace('\'', "")
    else:
        endereco = ""

    if row[3] is not None:
        row3_4 = row[3].split('-')
        cidade = row3_4[0].title().strip().replace('\'', "")
        estado = row3_4[1].upper().strip().replace('\'', "")
    else:
        cidade = ""
        estado = ""
    
    if row[4] is not None:
        cep = ''.join(str(row[4]).split('-')).strip()
    else:
        cep = ""
    
    if row[5] is not None:
        if type(row[5]) is str:
            try:
                data_nascimento = datetime.datetime.strptime(row[5], "%Y-%m-%d %H:%M:%S")
            except ValueError:
                data_nascimento = "Null"
        else:
            data_nascimento = row[5]
    else:
        data_nascimento = "Null"

    file.write(insert_command)

    if data_nascimento != "Null":
        file.write(f"\t('{nome}', '{endereco}', '{cidade}', '{estado}', '{cep}', '{data_nascimento}', CURDATE());\n\n")
    else:
        file.write(f"\t('{nome}', '{endereco}', '{cidade}', '{estado}', '{cep}', {data_nascimento}, CURDATE());\n\n")

    if row[6] is not None:
        telefones = str(row[6]).replace('OU', '/').replace('ou', '/').replace('-', '').split('/')

        if len(telefones):
            file.write('SET @SolicitanteId = LAST_INSERT_ID();\n\n')
            

            count = 1

            for telefone in telefones:
                update = []

                if '(' not in telefone:
                    
                    if len(telefone) > 9:
                        telefone_observacao = telefone
                        telefone = "999999999"
                        file.write('INSERT INTO telefones(ddd, numero, solicitante_id) VALUES\n')
                        file.write(f"\t('24', '{telefone.strip()}', @SolicitanteId);\n")
                        update.append("Número de Telefone : " + telefone_observacao.strip())
                    else:
                        file.write('INSERT INTO telefones(ddd, numero, solicitante_id) VALUES\n')
                        file.write(f"\t('24', '{telefone.strip()}', @SolicitanteId);")
                
                else:
                    ddd = telefone.replace(')', '(')
                    ddd = ddd.split("(")
                    telefone = ddd[2]

                    if len(telefone) > 9:
                        telefone = "999999999"
                        file.write('INSERT INTO telefones(ddd, numero, solicitante_id) VALUES\n')
                        file.write(f"\t('{ddd[1].strip()}', '{telefone.strip()}', @SolicitanteId);\n")
                        update.append("Número de Telefone : " + ddd[2].strip())
                    else:
                        file.write('INSERT INTO telefones(ddd, numero, solicitante_id) VALUES\n')
                        file.write(f"\t('{ddd[1].strip()}', '{telefone.strip()}', @SolicitanteId);")
                
                

                
                file.write('\n')
               

                count += 1
            if update:
                file.write(f"UPDATE solicitantes SET observacao = {update[0]} WHERE solicitante_id = @SolicitanteId;\n")

os.startfile("output.txt") 