from openpyxl import load_workbook
import os

try:
    os.remove('output.txt')
except OSError:
    print('New file')

file = open('output.txt', 'w')

temp_table_name = '@Oportunidades'
insert_command = f'INSERT INTO {temp_table_name} (CodMec, UnidadeNegocio, TipoEstudo, IdEmpresa, ValorAlunado, AlunosEI, AlunosEF1, AlunosEF2, AlunosEM, AlunosPrevest, Acao, IdProposta) VALUES\n'

workbook = load_workbook('./CargaOfertaRecomendada.xlsx', read_only=True)
sheets_names = workbook.sheetnames

# expecting only 'oferta recomendadas'
for sheet_name in sheets_names:
    sheet = workbook[sheet_name]
    index = 1

    file.write(insert_command)

    for row in sheet.iter_rows(min_row=1, values_only=True):
        if index == 1000:
            file.write(f'\t({row[0]}, {row[1]}, {row[2]}, {row[3]}, {row[4]}, {row[5]}, {row[6]}, {row[7]}, {row[8]}, {row[9]}, \'{row[10]}\', {row[11]});\n\n')
            file.write(insert_command)
            index = 0
        else:
            file.write(f'\t({row[0]}, {row[1]}, {row[2]}, {row[3]}, {row[4]}, {row[5]}, {row[6]}, {row[7]}, {row[8]}, {row[9]}, \'{row[10]}\', {row[11]}),\n')

        index += 1